#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Агент: Підготовка таблиці друку дипломів та подяк
Фестиваль Toronto (toronto.org.ua) → Типографія Визитка
ТЗ версія 8.0 | Лютий 2026

Використання:
    python agent_druk.py \
        --excel "Січень 2026 оплата та товаря.xls.xlsx" \
        --pdf-diplomy "ТАБЛИЦЯ...ДИПЛОМА...pdf" \
        --pdf-podyaky "ТАБЛИЦЯ...ПОДЯКИ...pdf" \
        [--output "друк_01_2026.xlsx"] \
        [--month "Січень 2026"] \
        [--bitrix-url "https://..."]
"""

import argparse
import sys
import os
import re
import time
from datetime import datetime
from difflib import SequenceMatcher

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ Потрібно: pip install openpyxl")
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    print("❌ Потрібно: pip install pdfplumber")
    sys.exit(1)

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

# ==========================================
# КОНФІГУРАЦІЯ ЗА ЗАМОВЧУВАННЯМ
# ==========================================
DEFAULT_CONFIG = {
    'DIPLOMA_FIELD_ID':  'UF_CRM_DIPLOMA_NUMBER',
    'PODYAKA_FIELD_ID':  'UF_CRM_PODYAKA_NUMBER',
    'SHEET_DIPLOMY':     'Друк дипломів',
    'SHEET_PODYAKY':     'Друк подяк',
    'API_DELAY_MS':      200,
    'FUZZY_THRESHOLD':   0.75,
    'BITRIX_WEBHOOK_URL': '',
}

# Кольори форматування
COLOR_HEADER  = "4472C4"   # синій
COLOR_DIPLOMA = "DEEAF1"   # світло-блакитний
COLOR_PODYAKA = "E2EFDA"   # світло-зелений
COLOR_WARNING = "FFEB9C"   # жовтий


# ==========================================
# КЛАСИФІКАЦІЯ ТОВАРІВ  (розділ 3 ТЗ)
# ==========================================
def classify_product(product_name: str) -> str:
    """
    Повертає тип товару:
      'FULL'       – 590 грн повний комплект нагород
      'DIPLOMA'    – Диплом в друкованому вигляді (90 грн)
      'PODYAKA'    – Подяка керівнику в друковному вигляді (90 грн)
      'ELECTRONIC' – тільки електронні версії (190 грн)
      'OTHER'      – невідомий товар
    """
    if not product_name:
        return 'OTHER'
    p = str(product_name)
    if '590 грн' in p and 'повний комплект' in p:
        return 'FULL'
    if 'Диплом' in p and 'друкован' in p:
        return 'DIPLOMA'
    if 'Подяка керівнику' in p:
        return 'PODYAKA'
    if 'електронні версії' in p or 'тільки електронн' in p:
        return 'ELECTRONIC'
    return 'OTHER'


# ==========================================
# ЗЧИТУВАННЯ EXCEL
# ==========================================
def get_field(row: dict, *names, default=None):
    """Шукає поле у рядку за кількома можливими назвами колонок."""
    for name in names:
        if name in row and row[name] is not None and str(row[name]).strip() != '':
            return row[name]
    # нечіткий пошук по ключах
    names_lower = [n.lower() for n in names]
    for key in row:
        if key is None:
            continue
        k = str(key).lower()
        for n in names_lower:
            if n in k or k in n:
                if row[key] is not None and str(row[key]).strip() != '':
                    return row[key]
    return default


def _read_ws(ws) -> (list, list):
    """Читає аркуш openpyxl, повертає (headers, list_of_dicts)."""
    headers = None
    rows = []
    for row_vals in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row_vals):
            continue
        if headers is None:
            headers = [str(v).strip() if v is not None else f'_col{i}'
                       for i, v in enumerate(row_vals)]
            continue
        rows.append(dict(zip(headers, row_vals)))
    return headers, rows


def read_excel(path: str, config: dict) -> (list, list):
    """
    Читає Excel-файл.
    - Якщо аркуші «Друк дипломів» / «Друк подяк» існують → читає їх.
    - Інакше → фільтрує з першого аркуша за типом товару.
    """
    wb = openpyxl.load_workbook(path)
    sheet_d = config['SHEET_DIPLOMY']
    sheet_p = config['SHEET_PODYAKY']

    if sheet_d in wb.sheetnames and sheet_p in wb.sheetnames:
        print(f"📄 Знайдено аркуші: '{sheet_d}' і '{sheet_p}'")
        _, diplomy = _read_ws(wb[sheet_d])
        _, podyaky = _read_ws(wb[sheet_p])
        return diplomy, podyaky

    # Автофільтрація з головного аркуша
    main_name = wb.sheetnames[0]
    print(f"⚠  Аркуші не знайдено — фільтрую з '{main_name}'")
    _, all_rows = _read_ws(wb[main_name])

    diplomy, podyaky = [], []
    for row in all_rows:
        ptype = classify_product(str(get_field(row, 'Товар') or ''))
        if ptype in ('FULL', 'DIPLOMA'):
            diplomy.append(row)
        if ptype in ('FULL', 'PODYAKA'):
            podyaky.append(row)

    print(f"   → Відфільтровано: {len(diplomy)} дипломів, {len(podyaky)} подяк")
    return diplomy, podyaky


# ==========================================
# ЗЧИТУВАННЯ PDF
# ==========================================
def read_pdf_diplomy(path: str) -> list:
    """
    Читає PDF-таблицю номерів дипломів.
    Структура колонок: ID | Artist | Номінація | Назва роботи | Laureate | №Диплому
    Повертає list[dict]: id, artist, laureate, num_diploma
    """
    records = []
    print(f"📖 PDF дипломів: {os.path.basename(path)}")
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in table:
                    if not row or len(row) < 6:
                        continue
                    id_str  = str(row[0]).replace('\n', '').strip() if row[0] else ''
                    num_str = str(row[5]).replace('\n', '').strip() if row[5] else ''
                    if not id_str.isdigit() or not num_str.isdigit():
                        continue
                    records.append({
                        'id':          int(id_str),
                        'artist':      str(row[1]).replace('\n', ' ').strip() if row[1] else '',
                        'laureate':    str(row[4]).replace('\n', ' ').strip() if row[4] else '',
                        'num_diploma': int(num_str),
                    })
    print(f"   → {len(records)} записів")
    return records


def read_pdf_podyaky(path: str) -> list:
    """
    Читає PDF-таблицю номерів подяк.
    Структура колонок: ID | ПІБ керівника | №Подяки
    Повертає list[dict]: id, pib_kerivnyk, num_podyaka
    """
    records = []
    print(f"📖 PDF подяк: {os.path.basename(path)}")
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in table:
                    if not row or len(row) < 3:
                        continue
                    id_str  = str(row[0]).replace('\n', '').strip() if row[0] else ''
                    num_str = str(row[2]).replace('\n', '').strip() if row[2] else ''
                    if not id_str.isdigit() or not num_str.isdigit():
                        continue
                    records.append({
                        'id':           int(id_str),
                        'pib_kerivnyk': str(row[1]).replace('\n', ' ').strip() if row[1] else '',
                        'num_podyaka':  int(num_str),
                    })
    print(f"   → {len(records)} записів")
    return records


# ==========================================
# НЕЧІТКИЙ ПОШУК ПІБ
# ==========================================
def _norm(s: str) -> str:
    return re.sub(r'\s+', ' ', str(s).lower().strip())


def fuzzy_match(a: str, b: str, threshold: float = 0.75) -> bool:
    a, b = _norm(a), _norm(b)
    if not a or not b:
        return False
    if a == b or a in b or b in a:
        return True
    return SequenceMatcher(None, a, b).ratio() >= threshold


def find_podyaka(pib_kerivnyk: str, podyaky_pdf: list, threshold: float = 0.75):
    """
    Шукає номер подяки за ПІБ керівника (fuzzy match).
    Якщо ПІБ містить кількох осіб (через кому) — шукає по кожному.
    Повертає (num_podyaka, matched_pib) або (None, None).
    """
    if not pib_kerivnyk:
        return None, None
    pib_str = str(pib_kerivnyk)

    for rec in podyaky_pdf:
        if fuzzy_match(pib_str, rec['pib_kerivnyk'], threshold):
            return rec['num_podyaka'], rec['pib_kerivnyk']

    # Якщо кілька ПІБ через кому
    for part in [p.strip() for p in pib_str.split(',')]:
        if len(part) < 5:
            continue
        for rec in podyaky_pdf:
            if fuzzy_match(part, rec['pib_kerivnyk'], threshold):
                return rec['num_podyaka'], rec['pib_kerivnyk']

    return None, None


def clean_participant_pib(pib: str) -> str:
    """Очищує ПІБ учасника: видаляє англійський переклад після ' - '."""
    if not pib:
        return ''
    return str(pib).split(' - ')[0].strip()


def clean_teacher_pib(pib: str) -> str:
    """
    Очищує ПІБ керівника:
    1. Видаляє слово 'керівник'
    2. Видаляє англійський переклад після ' - '
    3. Залишає тільки частини через кому, що схожі на ПІБ
       (кожна частина: 2+ слів, кожне з великої літери)
    """
    if not pib:
        return ''
    # 1. Видаляємо 'керівник'
    result = re.sub(r'\bкерівник\w*\b', '', str(pib), flags=re.IGNORECASE)
    # 2. Видаляємо англійський переклад
    result = result.split(' - ')[0]
    # 3. Фільтруємо частини через кому
    parts = [p.strip() for p in result.split(',') if p.strip()]
    pib_parts = []
    for part in parts:
        words = part.split()
        # ПІБ: 2–4 слова, кожне починається з великої літери
        if len(words) >= 2 and all(w and w[0].isupper() for w in words):
            pib_parts.append(part)
    result = ', '.join(pib_parts) if pib_parts else (parts[0] if parts else '')
    return re.sub(r'\s+', ' ', result).strip(' ,;')


# ==========================================
# КРОК 2 – ОБРОБКА АРКУША «ДРУК ДИПЛОМІВ»
# ==========================================
def process_diplomy(diplomy_rows: list, diplomy_pdf: list, podyaky_pdf: list,
                    config: dict, errors: list) -> (list, list):
    """
    Для кожного рядка аркуша 'Друк дипломів':
      - Знаходить номер(и) диплому в PDF за ID.
      - Якщо 590 грн (FULL) – також знаходить №Подяки за ПІБ керівника.
    Повертає (diploma_out, podyaka_out).
    """
    diploma_out  = []
    podyaka_out  = []
    threshold    = config.get('FUZZY_THRESHOLD', 0.75)

    # Індекс PDF дипломів: id → list[rec]
    pdf_by_id = {}
    for rec in diplomy_pdf:
        pdf_by_id.setdefault(rec['id'], []).append(rec)

    for row in diplomy_rows:
        raw_id = get_field(row, 'ID')
        if raw_id is None:
            continue
        try:
            deal_id = int(raw_id)
        except (ValueError, TypeError):
            continue

        pib_u    = clean_participant_pib(str(get_field(row, 'ПІБ Учасника', 'ПІБ учасника', 'Artist') or ''))
        pib_k    = str(get_field(row, 'ПІБ керівника, концертмейстера',
                                      'ПІБ керівника') or '')
        product  = str(get_field(row, 'Товар') or '')
        qty_raw  = get_field(row, 'Кількість')
        qty      = int(qty_raw) if qty_raw is not None else 1
        ptype    = classify_product(product)

        if ptype == 'ELECTRONIC':
            continue

        # --- Диплом ---
        pdf_recs = pdf_by_id.get(deal_id, [])
        if not pdf_recs:
            diploma_out.append({'num_doc': '⚠ Не знайдено', 'type': 'Диплом',
                                'pib': pib_u, 'qty': qty, 'id': deal_id, 'warning': True})
            errors.append(f'ID {deal_id} ({pib_u}): номер диплому не знайдено в PDF')
        else:
            for pdf_rec in pdf_recs:
                diploma_out.append({'num_doc': pdf_rec['num_diploma'], 'type': 'Диплом',
                                    'pib': pib_u, 'qty': qty, 'id': deal_id, 'warning': False})

        # --- Подяка (тільки для 590 грн) ---
        # qty для подяки = 1: навіть якщо учасників 2+, вчитель отримує 1 примірник
        if ptype == 'FULL':
            num_pod, _ = find_podyaka(pib_k, podyaky_pdf, threshold)
            if num_pod is None:
                podyaka_out.append({'num_doc': '⚠ Не знайдено', 'type': 'Подяка',
                                    'pib': clean_teacher_pib(pib_k), 'qty': 1,
                                    'id': deal_id, 'warning': True})
                errors.append(f'ID {deal_id}: подяка для «{pib_k}» не знайдена')
            else:
                podyaka_out.append({'num_doc': num_pod, 'type': 'Подяка',
                                    'pib': clean_teacher_pib(pib_k), 'qty': 1,
                                    'id': deal_id, 'warning': False})

    return diploma_out, podyaka_out


# ==========================================
# КРОК 3 – ОБРОБКА АРКУША «ДРУК ПОДЯК»
# ==========================================
def process_podyaky(podyaky_rows: list, podyaky_pdf: list,
                    config: dict, errors: list) -> list:
    """
    Для кожного рядка аркуша 'Друк подяк':
      - Якщо FULL (590 грн) – пропускаємо (вже оброблено в кроці 2).
      - Якщо PODYAKA (90 грн) – знаходить №Подяки за ПІБ керівника.
    Повертає podyaka_out.
    """
    podyaka_out = []
    threshold   = config.get('FUZZY_THRESHOLD', 0.75)

    for row in podyaky_rows:
        raw_id = get_field(row, 'ID')
        if raw_id is None:
            continue
        try:
            deal_id = int(raw_id)
        except (ValueError, TypeError):
            continue

        pib_k   = str(get_field(row, 'ПІБ керівника, концертмейстера',
                                      'ПІБ керівника') or '')
        product = str(get_field(row, 'Товар') or '')
        qty_raw = get_field(row, 'Кількість')
        qty     = int(qty_raw) if qty_raw is not None else 1
        ptype   = classify_product(product)

        if ptype == 'ELECTRONIC':
            continue
        if ptype == 'FULL':
            continue   # вже оброблено в кроці 2

        num_pod, _ = find_podyaka(pib_k, podyaky_pdf, threshold)
        if num_pod is None:
            podyaka_out.append({'num_doc': '⚠ Не знайдено', 'type': 'Подяка',
                                'pib': clean_teacher_pib(pib_k), 'qty': qty,
                                'id': deal_id, 'warning': True})
            errors.append(f'ID {deal_id}: подяка для «{pib_k}» не знайдена')
        else:
            podyaka_out.append({'num_doc': num_pod, 'type': 'Подяка',
                                'pib': clean_teacher_pib(pib_k), 'qty': qty,
                                'id': deal_id, 'warning': False})

    return podyaka_out


# ==========================================
# КРОК 4 – ФОРМУВАННЯ ЗВЕДЕНОЇ ТАБЛИЦІ
# ==========================================
def build_zvedena(diploma_out: list, podyaka_out_all: list,
                  diplomy_rows: list) -> list:
    """
    Будує зведену таблицю.
    Порядок: для кожного диплому → одразу всі подяки вчителя (якщо не ще не виводились).
    Це забезпечує групування подяк одного вчителя разом.
    """
    # Карта: deal_id → normalized teacher PIB
    id_to_teacher = {}
    for row in diplomy_rows:
        raw = get_field(row, 'ID')
        if raw is None:
            continue
        try:
            deal_id = int(raw)
        except:
            continue
        pib_k = get_field(row, 'ПІБ керівника, концертмейстера', 'ПІБ керівника')
        if pib_k:
            id_to_teacher[deal_id] = _norm(str(pib_k))

    # Карта: normalized teacher → list of podyaka records
    teacher_to_podyakas = {}
    for rec in podyaka_out_all:
        teacher = id_to_teacher.get(rec['id'], _norm(rec['pib']))
        teacher_to_podyakas.setdefault(teacher, []).append(rec)

    result           = []
    emitted_teachers = set()

    for diploma in diploma_out:
        result.append(diploma)
        teacher = id_to_teacher.get(diploma['id'])
        if teacher and teacher not in emitted_teachers:
            podyakas = teacher_to_podyakas.get(teacher, [])
            if podyakas:
                result.extend(podyakas)
                emitted_teachers.add(teacher)

    # Подяки, що залишились не виведеними
    for rec in podyaka_out_all:
        teacher = id_to_teacher.get(rec['id'], _norm(rec['pib']))
        if teacher not in emitted_teachers:
            result.append(rec)
            emitted_teachers.add(teacher)

    return result


# ==========================================
# ЗАПИС ВИХІДНОГО XLSX
# ==========================================
def _hdr(ws, row_num: int):
    fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    font = Font(bold=True, color='FFFFFF')
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')


def _warn_fill():
    return PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type='solid')


def write_output(diploma_out: list, podyaka_out: list, zvedena: list,
                 output_path: str, month: str, errors: list):
    """Записує три аркуші вихідного Excel."""
    wb = openpyxl.Workbook()

    # ── Аркуш 1: Друк дипломів ────────────────────────────────
    ws_d = wb.active
    ws_d.title = 'Друк дипломів'
    ws_d.append([f'ТАБЛИЦЯ ДРУКУ ДИПЛОМІВ — Типографія Визитка | {month}',
                 None, None, None, None])
    ws_d.merge_cells('A1:E1')
    ws_d['A1'].font = Font(bold=True, size=12)
    ws_d['A1'].alignment = Alignment(horizontal='center')
    ws_d.append(['№', '№ Диплому', 'ПІБ учасника', 'Кількість', 'ID угоди'])
    _hdr(ws_d, 2)

    warn_d = []
    for i, rec in enumerate(diploma_out, 1):
        ws_d.append([i, rec['num_doc'], rec['pib'], rec['qty'], rec['id']])
        if rec.get('warning'):
            warn_d.append(i + 2)

    last_d = len(diploma_out) + 2
    ws_d.append(['Всього дипломів до друку (без помилок):', None, None,
                 f'=SUMIF(B3:B{last_d},"<>*Не знайдено*",D3:D{last_d})', None])

    if warn_d:
        ids_warn = [diploma_out[r - 3]['id'] for r in warn_d]
        names_warn = [diploma_out[r - 3]['pib'] for r in warn_d]
        parts = ', '.join(f"ID {i} ({n})" for i, n in zip(ids_warn, names_warn))
        ws_d.append([f'⚠ Рядки {", ".join(str(r-2) for r in warn_d)} — номер диплому не знайдено в PDF. '
                     f'Необхідно перевірити: {parts}.', None, None, None, None])

    for r in warn_d:
        for cell in ws_d[r]:
            cell.fill = _warn_fill()

    for col, w in zip('ABCDE', [5, 15, 45, 12, 10]):
        ws_d.column_dimensions[col].width = w

    # ── Аркуш 2: Друк подяк ───────────────────────────────────
    ws_p = wb.create_sheet('Друк подяк')
    ws_p.append([f'ТАБЛИЦЯ ДРУКУ ПОДЯК — Типографія Визитка | {month}',
                 None, None, None, None])
    ws_p.merge_cells('A1:E1')
    ws_p['A1'].font = Font(bold=True, size=12)
    ws_p['A1'].alignment = Alignment(horizontal='center')
    ws_p.append(['№', '№ Подяки', 'ПІБ керівника', 'Кількість', 'ID угоди'])
    _hdr(ws_p, 2)

    warn_p = []
    for i, rec in enumerate(podyaka_out, 1):
        ws_p.append([i, rec['num_doc'], rec['pib'], rec['qty'], str(rec['id'])])
        if rec.get('warning'):
            warn_p.append(i + 2)

    last_p = len(podyaka_out) + 2
    ws_p.append(['Всього подяк до друку:', None, None,
                 f'=SUM(D3:D{last_p})', None])

    for r in warn_p:
        for cell in ws_p[r]:
            cell.fill = _warn_fill()

    for col, w in zip('ABCDE', [5, 12, 45, 12, 10]):
        ws_p.column_dimensions[col].width = w

    # ── Аркуш 3: Зведена для типографії ──────────────────────
    ws_z = wb.create_sheet('Зведена для типографії')
    ws_z.append([f'ЗВЕДЕНА ТАБЛИЦЯ ДРУКУ — Типографія Визитка | {month}',
                 None, None, None, None, None])
    ws_z.merge_cells('A1:F1')
    ws_z['A1'].font = Font(bold=True, size=12)
    ws_z['A1'].alignment = Alignment(horizontal='center')
    ws_z.append(['№', '№ Документу', 'Тип', 'ПІБ', 'К-сть', 'ID угоди'])
    _hdr(ws_z, 2)

    fill_d = PatternFill(start_color=COLOR_DIPLOMA, end_color=COLOR_DIPLOMA, fill_type='solid')
    fill_p = PatternFill(start_color=COLOR_PODYAKA, end_color=COLOR_PODYAKA, fill_type='solid')

    for i, rec in enumerate(zvedena, 1):
        ws_z.append([i, rec['num_doc'], rec['type'], rec['pib'],
                     rec['qty'], str(rec['id'])])
        fill = _warn_fill() if rec.get('warning') else \
               (fill_d if rec['type'] == 'Диплом' else fill_p)
        for cell in ws_z[i + 2]:
            cell.fill = fill

    last_z = len(zvedena) + 2
    ws_z.append(['ВСЬОГО до друку (без помилок):', None, None, None,
                 f'=SUMIF(B3:B{last_z},"<>*Не знайдено*",E3:E{last_z})', None])

    for col, w in zip('ABCDEF', [5, 15, 10, 45, 8, 10]):
        ws_z.column_dimensions[col].width = w

    wb.save(output_path)
    print(f"\n✅ Збережено: {output_path}")


# ==========================================
# КРОК 5 – БІТРІКС REST API
# ==========================================
def update_bitrix(diploma_out: list, podyaka_out: list, config: dict, errors: list):
    webhook = config.get('BITRIX_WEBHOOK_URL', '').strip()
    if not webhook:
        print("⚠  BITRIX_WEBHOOK_URL не задано — пропускаю API")
        return
    if not REQUESTS_AVAILABLE:
        print("⚠  pip install requests — пропускаю API")
        return

    d_field = config['DIPLOMA_FIELD_ID']
    p_field = config['PODYAKA_FIELD_ID']
    delay   = config['API_DELAY_MS'] / 1000.0

    diploma_by_id  = {}
    for rec in diploma_out:
        diploma_by_id.setdefault(rec['id'], []).append(rec)
    podyaka_by_id  = {}
    for rec in podyaka_out:
        podyaka_by_id.setdefault(rec['id'], []).append(rec)

    all_ids = set(diploma_by_id) | set(podyaka_by_id)
    for deal_id in sorted(all_ids):
        fields = {}
        d_nums = [str(r['num_doc']) for r in diploma_by_id.get(deal_id, []) if not r.get('warning')]
        if d_nums:
            fields[d_field] = ', '.join(d_nums)
        p_nums = list({str(r['num_doc']) for r in podyaka_by_id.get(deal_id, []) if not r.get('warning')})
        if p_nums:
            fields[p_field] = ', '.join(p_nums)
        if not fields:
            continue

        url = webhook.rstrip('/') + '/crm.deal.update.json'
        try:
            resp = requests.post(url, json={'id': deal_id, 'fields': fields}, timeout=10)
            if resp.ok and resp.json().get('result'):
                print(f"  ✓ {deal_id}: {fields}")
            else:
                msg = resp.text[:200]
                print(f"  ✗ {deal_id}: {msg}")
                errors.append(f'API помилка ID {deal_id}: {msg}')
        except Exception as e:
            print(f"  ✗ {deal_id}: {e}")
            errors.append(f'API виняток ID {deal_id}: {e}')
        time.sleep(delay)


# ==========================================
# ЗВІТ ПОМИЛОК
# ==========================================
def write_errors(errors: list, out_dir: str):
    if not errors:
        return
    path = os.path.join(out_dir or '.', 'звіт_помилок.txt')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(f"Звіт помилок — {datetime.now():%Y-%m-%d %H:%M}\n{'='*60}\n")
        for e in errors:
            f.write(f"• {e}\n")
    print(f"⚠  Звіт помилок: {path} ({len(errors)} помилок)")


# ==========================================
# ПОРІВНЯННЯ З ЕТАЛОНОМ (для тестування)
# ==========================================
def compare_with_reference(result_path: str, reference_path: str):
    """Порівнює вихідний файл з еталонним і виводить різниці."""
    wb_r = openpyxl.load_workbook(result_path)
    wb_e = openpyxl.load_workbook(reference_path)

    print(f"\n{'='*60}")
    print("  ПОРІВНЯННЯ З ЕТАЛОНОМ")
    print(f"{'='*60}")

    for sheet_name in ['Друк дипломів', 'Друк подяк', 'Зведена для типографії']:
        if sheet_name not in wb_r.sheetnames:
            print(f"  ⚠ Аркуш '{sheet_name}' відсутній у результаті")
            continue
        if sheet_name not in wb_e.sheetnames:
            print(f"  ⚠ Аркуш '{sheet_name}' відсутній в еталоні")
            continue

        ws_r = wb_r[sheet_name]
        ws_e = wb_e[sheet_name]

        rows_r = [row for row in ws_r.iter_rows(values_only=True)
                  if any(c is not None for c in row)]
        rows_e = [row for row in ws_e.iter_rows(values_only=True)
                  if any(c is not None for c in row)]

        # Беремо рядки даних (без заголовку і підсумків)
        data_r = []
        data_e = []
        header_found = False
        for row in rows_r:
            vals = [str(v) if v is not None else '' for v in row]
            if '№ Диплому' in vals or '№ Подяки' in vals or '№ Документу' in vals:
                header_found = True
                continue
            if header_found and vals[0].isdigit() if vals else False:
                data_r.append(tuple(str(v) if v is not None else '' for v in row))

        header_found = False
        for row in rows_e:
            vals = [str(v) if v is not None else '' for v in row]
            if '№ Диплому' in vals or '№ Подяки' in vals or '№ Документу' in vals:
                header_found = True
                continue
            if header_found and vals[0].isdigit() if vals else False:
                data_e.append(tuple(str(v) if v is not None else '' for v in row))

        print(f"\n─── {sheet_name} ───")
        print(f"  Еталон: {len(data_e)} рядків  |  Результат: {len(data_r)} рядків")

        ok = 0
        diff = 0
        for i, (r, e) in enumerate(zip(data_r, data_e), 1):
            # Порівнюємо ключові поля: №Документу (col 1), Тип/ПІБ (col 2/3), К-сть
            if r[1:5] == e[1:5]:
                ok += 1
            else:
                diff += 1
                print(f"  ✗ Рядок {i}:")
                print(f"      Еталон:    {e}")
                print(f"      Результат: {r}")

        if len(data_r) != len(data_e):
            extra = abs(len(data_r) - len(data_e))
            print(f"  ⚠ Різниця у кількості рядків: {extra}")

        print(f"  ✓ Співпадають: {ok}  |  ✗ Різниці: {diff}")

    print(f"\n{'='*60}\n")


# ==========================================
# ТОЧКА ВХОДУ
# ==========================================
def main():
    parser = argparse.ArgumentParser(
        description='Агент друку дипломів та подяк — Типографія Визитка'
    )
    parser.add_argument('--excel',       required=True)
    parser.add_argument('--pdf-diplomy', required=True)
    parser.add_argument('--pdf-podyaky', required=True)
    parser.add_argument('--output',      default=None)
    parser.add_argument('--month',       default=None)
    parser.add_argument('--bitrix-url',  default='')
    parser.add_argument('--no-api',      action='store_true')
    parser.add_argument('--compare',     default=None,
                        help='Шлях до еталонного файлу для порівняння')
    args = parser.parse_args()

    month = args.month or datetime.now().strftime('%B %Y')
    if args.output:
        output_path = args.output
    else:
        now = datetime.now()
        output_path = f"друк_{now.strftime('%m_%Y')}.xlsx"

    config = {**DEFAULT_CONFIG}
    if args.bitrix_url:
        config['BITRIX_WEBHOOK_URL'] = args.bitrix_url

    print(f"\n{'='*60}\n  Агент друку дипломів — {month}\n{'='*60}\n")
    errors = []

    # Крок 1: читання
    print("─── Крок 1: Читання ───")
    try:
        diplomy_rows, podyaky_rows = read_excel(args.excel, config)
    except Exception as e:
        print(f"❌ {e}"); sys.exit(1)

    diplomy_pdf = read_pdf_diplomy(args.pdf_diplomy)
    podyaky_pdf = read_pdf_podyaky(args.pdf_podyaky)

    print(f"\n  Рядків дипломів: {len(diplomy_rows)}")
    print(f"  Рядків подяк:    {len(podyaky_rows)}")
    print(f"  PDF дипломів:    {len(diplomy_pdf)}")
    print(f"  PDF подяк:       {len(podyaky_pdf)}")

    # Крок 2
    print("\n─── Крок 2: Обробка 'Друк дипломів' ───")
    diploma_out, podyaka_from_d = process_diplomy(
        diplomy_rows, diplomy_pdf, podyaky_pdf, config, errors)
    print(f"  Дипломів: {len(diploma_out)}  |  Подяк (590 грн): {len(podyaka_from_d)}")

    # Крок 3
    print("\n─── Крок 3: Обробка 'Друк подяк' ───")
    podyaka_from_p = process_podyaky(
        podyaky_rows, podyaky_pdf, config, errors)
    print(f"  Подяк (90 грн): {len(podyaka_from_p)}")

    # Об'єднуємо та сортуємо подяки за ID (спадаючий) — для відповідності порядку еталону
    podyaka_out_all = sorted(
        podyaka_from_d + podyaka_from_p,
        key=lambda r: r['id'],
        reverse=True
    )
    print(f"\n  РАЗОМ дипломів: {len(diploma_out)}  |  подяк: {len(podyaka_out_all)}")

    # Крок 4
    print("\n─── Крок 4: Формування зведеної таблиці ───")
    zvedena = build_zvedena(diploma_out, podyaka_out_all, diplomy_rows)
    print(f"  Рядків у зведеній: {len(zvedena)}")

    write_output(diploma_out, podyaka_out_all, zvedena,
                 output_path, month, errors)

    # Крок 5
    if not args.no_api:
        print("\n─── Крок 5: Бітрікс API ───")
        update_bitrix(diploma_out, podyaka_out_all, config, errors)

    if errors:
        write_errors(errors, os.path.dirname(output_path) or '.')
        print(f"\n⚠  {len(errors)} помилок — дивіться звіт_помилок.txt")
    else:
        print("\n✅ Помилок не виявлено!")

    # Порівняння з еталоном
    if args.compare:
        compare_with_reference(output_path, args.compare)

    print(f"\n{'='*60}\n  Готово! → {output_path}\n{'='*60}\n")
    return 0


if __name__ == '__main__':
    sys.exit(main())
